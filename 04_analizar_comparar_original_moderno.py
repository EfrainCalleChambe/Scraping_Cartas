# -*- coding: utf-8 -*-
"""
04_ANALIZAR_COMPARAR_ORIGINAL_MODERNO_CORREGIDO.py

Versión corregida:
- No genera CSV.
- No crea carpetas CSV.
- Busca automáticamente las carpetas aunque estén dentro de otra carpeta.
- Exporta todo lo tabular a Excel (.xlsx).
- Genera gráficos PNG, redes PNG, nubes PNG y HTML comparativo.

REQUISITOS:
    pip install pandas openpyxl matplotlib scikit-learn regex wordcloud networkx

EJECUTAR:
    python 04_ANALIZAR_COMPARAR_ORIGINAL_MODERNO_CORREGIDO.py
"""

from __future__ import annotations

import re
import math
import html
import difflib
import shutil
from pathlib import Path
from collections import Counter

import pandas as pd
import matplotlib.pyplot as plt

try:
    import regex as regex_unicode
    TIENE_REGEX = True
except Exception:
    regex_unicode = None
    TIENE_REGEX = False

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    TIENE_SKLEARN = True
except Exception:
    TIENE_SKLEARN = False

try:
    from wordcloud import WordCloud
    TIENE_WORDCLOUD = True
except Exception:
    TIENE_WORDCLOUD = False

try:
    import networkx as nx
    TIENE_NETWORKX = True
except Exception:
    TIENE_NETWORKX = False


# ============================================================
# CONFIGURACIÓN
# ============================================================

# Carpeta base donde estás ejecutando el script.
# Normalmente déjalo así.
BASE_DIR = Path(".").resolve()

# Si quieres forzar rutas exactas, coloca aquí las rutas.
# Ejemplo:
# CARPETA_ORIGINAL_TXT = Path(r"C:\Users\dnce72\Documents\FRAYO\Scraping\montemar_transcripciones\txt_por_pagina")
# CARPETA_MODERNA_TXT = Path(r"C:\Users\dnce72\Documents\FRAYO\Scraping\montemar_moderno_chatgpt_todos_lotes_hibrido\txt_moderno_por_pagina")
CARPETA_ORIGINAL_TXT = None
CARPETA_MODERNA_TXT = None

# Nombres probables de carpetas.
NOMBRES_ORIGINAL = [
    "montemar_transcripciones/txt_por_pagina",
    "montemar_transcripciones_extracted/montemar_transcripciones/txt_por_pagina",
]

NOMBRES_MODERNO = [
    "montemar_moderno_chatgpt_todos_lotes_hibrido/txt_moderno_por_pagina",
    "montemar_moderno_chatgpt_lote01/txt_moderno_por_pagina",
    "montemar_moderno_gemini/txt_moderno_por_pagina",
    "montemar_moderno_ollama/txt_moderno_por_pagina",
]

# Salidas.
CARPETA_ANALISIS_MODERNO = Path("analisis_moderno_montemar")
CARPETA_MODERNO_GRAFICOS = CARPETA_ANALISIS_MODERNO / "graficos"
CARPETA_MODERNO_NUBES = CARPETA_ANALISIS_MODERNO / "nubes_palabras"
CARPETA_MODERNO_REDES = CARPETA_ANALISIS_MODERNO / "redes_asociaciones"
CARPETA_REEMPLAZOS = CARPETA_ANALISIS_MODERNO / "reemplazos"
EXCEL_MODERNO = CARPETA_ANALISIS_MODERNO / "analisis_solo_moderno_montemar.xlsx"

CARPETA_COMPARACION = Path("analisis_comparacion_montemar")
CARPETA_COMP_GRAFICOS = CARPETA_COMPARACION / "graficos"
CARPETA_COMP_HTML = CARPETA_COMPARACION / "comparacion_html_por_pagina"
EXCEL_COMPARACION = CARPETA_COMPARACION / "analisis_comparacion_original_moderno.xlsx"

PASAR_A_MINUSCULAS = True
USAR_STOPWORDS = False
QUITAR_STOPWORDS_EN_REDES = True

TOP_FRECUENCIAS_PAGINA = 80
TOP_FRECUENCIAS_CARTA = 150
TOP_TFIDF_CARTA = 60
TOP_ASOCIACIONES_CARTA = 150
TOP_CAMBIOS_CARTA = 120
TOP_VOCAB_EXCLUSIVO = 300

VENTANA_ASOCIACION = 4

GENERAR_NUBES = True
TOP_NUBES_CARTAS = 25

GENERAR_REDES_ASOCIACIONES = True
TOP_NODOS_RED_CARTA = 45
TOP_ARISTAS_RED_CARTA = 70
TOP_NODOS_RED_GLOBAL = 70
TOP_ARISTAS_RED_GLOBAL = 120
MIN_PESO_ARISTA = 2

TOP_REEMPLAZOS_GLOBAL = 500
TOP_REEMPLAZOS_CARTA = 150
TOP_REEMPLAZOS_PAGINA = 100
GENERAR_RED_REEMPLAZOS = True
TOP_ARISTAS_RED_REEMPLAZOS = 120


STOPWORDS_ES = {
    "a", "al", "algo", "algunas", "algunos", "ante", "antes", "como", "con",
    "contra", "cual", "cuando", "de", "del", "desde", "donde", "durante",
    "e", "el", "ella", "ellas", "ellos", "en", "entre", "era", "eran",
    "eres", "es", "esa", "esas", "ese", "eso", "esos", "esta", "estaba",
    "estado", "estan", "están", "estar", "este", "esto", "estos", "fue",
    "fueron", "ha", "habia", "había", "han", "hasta", "hay", "la", "las",
    "le", "les", "lo", "los", "mas", "más", "me", "mi", "mis", "muy",
    "ni", "no", "nos", "o", "para", "pero", "por", "porque", "que", "qué",
    "se", "ser", "si", "sí", "sin", "sobre", "su", "sus", "te", "ti",
    "tu", "tus", "un", "una", "uno", "unos", "y", "ya", "yo",
    "vuestra", "merced"
}


# ============================================================
# CARPETAS
# ============================================================

def buscar_carpeta_por_patron(nombre_final: str, patron_archivo: str) -> Path | None:
    """
    Busca recursivamente una carpeta cuyo final termine en nombre_final
    y que contenga archivos que coincidan con patron_archivo.
    """
    for carpeta in BASE_DIR.rglob("*"):
        if not carpeta.is_dir():
            continue

        ruta_norm = str(carpeta).replace("\\", "/").lower()
        final_norm = nombre_final.replace("\\", "/").lower()

        if ruta_norm.endswith(final_norm):
            if list(carpeta.glob(patron_archivo)):
                return carpeta

    return None


def detectar_carpeta_original() -> Path:
    if CARPETA_ORIGINAL_TXT is not None:
        carpeta = Path(CARPETA_ORIGINAL_TXT)
        if carpeta.exists() and list(carpeta.glob("*_page_*.txt")):
            return carpeta
        raise FileNotFoundError(f"No existe o no tiene TXT originales: {carpeta}")

    for nombre in NOMBRES_ORIGINAL:
        carpeta = BASE_DIR / nombre
        if carpeta.exists() and list(carpeta.glob("*_page_*.txt")):
            return carpeta

    encontrada = buscar_carpeta_por_patron("txt_por_pagina", "*_page_*.txt")
    if encontrada:
        return encontrada

    raise FileNotFoundError(
        "No encontré la carpeta original txt_por_pagina.\n"
        "Debe existir algo como:\n"
        "  montemar_transcripciones/txt_por_pagina/\n\n"
        "Solución rápida: edita CARPETA_ORIGINAL_TXT al inicio del script con la ruta exacta."
    )


def detectar_carpeta_moderna() -> Path:
    if CARPETA_MODERNA_TXT is not None:
        carpeta = Path(CARPETA_MODERNA_TXT)
        if carpeta.exists() and list(carpeta.glob("*_moderno.txt")):
            return carpeta
        raise FileNotFoundError(f"No existe o no tiene TXT modernos: {carpeta}")

    for nombre in NOMBRES_MODERNO:
        carpeta = BASE_DIR / nombre
        if carpeta.exists() and list(carpeta.glob("*_moderno.txt")):
            return carpeta

    encontrada = buscar_carpeta_por_patron("txt_moderno_por_pagina", "*_moderno.txt")
    if encontrada:
        return encontrada

    raise FileNotFoundError(
        "No encontré la carpeta de TXT modernos.\n\n"
        "Debe existir algo como:\n"
        "  montemar_moderno_chatgpt_todos_lotes_hibrido/txt_moderno_por_pagina/\n\n"
        "Solución rápida:\n"
        "1. Descomprime el ZIP moderno en la misma carpeta donde ejecutas este script.\n"
        "2. Verifica que exista la carpeta txt_moderno_por_pagina.\n"
        "3. O edita CARPETA_MODERNA_TXT al inicio del script con la ruta exacta."
    )


def crear_carpetas() -> None:
    """
    Crea solo carpetas necesarias. No crea carpetas CSV.
    También elimina carpetas csv antiguas si existieran.
    """
    for vieja in [
        CARPETA_ANALISIS_MODERNO / "csv",
        CARPETA_COMPARACION / "csv",
    ]:
        if vieja.exists():
            shutil.rmtree(vieja)

    for c in [
        CARPETA_ANALISIS_MODERNO,
        CARPETA_MODERNO_GRAFICOS,
        CARPETA_MODERNO_NUBES,
        CARPETA_MODERNO_REDES,
        CARPETA_REEMPLAZOS,
        CARPETA_COMPARACION,
        CARPETA_COMP_GRAFICOS,
        CARPETA_COMP_HTML,
    ]:
        c.mkdir(parents=True, exist_ok=True)


# ============================================================
# UTILIDADES DE TEXTO
# ============================================================

def limpiar_nombre_archivo(s: str, max_len: int = 180) -> str:
    s = html.unescape(str(s))
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = re.sub(r"[\u0000-\u001f]+", "", s)
    return s[:max_len].strip(" ._")


def leer_texto(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace").strip()


def extraer_num_carta(carta_id: str):
    m = re.match(r"^(\d+)_", str(carta_id))
    return int(m.group(1)) if m else None


def parse_original(path: Path) -> dict:
    m = re.match(r"(.+)_page_(\d+)\.txt$", path.name, flags=re.I)
    if not m:
        return {}

    carta_id = m.group(1)
    pagina = int(m.group(2))

    return {
        "key": f"{carta_id}_page_{pagina:03d}",
        "carta_id": carta_id,
        "nro_carta": extraer_num_carta(carta_id),
        "pagina": pagina,
        "path_original": path,
    }


def parse_moderno(path: Path) -> dict:
    m = re.match(r"(.+)_page_(\d+)_moderno\.txt$", path.name, flags=re.I)
    if not m:
        return {}

    carta_id = m.group(1)
    pagina = int(m.group(2))

    return {
        "key": f"{carta_id}_page_{pagina:03d}",
        "carta_id": carta_id,
        "nro_carta": extraer_num_carta(carta_id),
        "pagina": pagina,
        "path_moderno": path,
    }


def normalizar_texto_base(texto: str) -> str:
    texto = html.unescape(str(texto))
    texto = texto.replace("\xa0", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def tokenizar(texto: str, quitar_stopwords: bool | None = None) -> list[str]:
    if quitar_stopwords is None:
        quitar_stopwords = USAR_STOPWORDS

    texto = normalizar_texto_base(texto)

    if PASAR_A_MINUSCULAS:
        texto = texto.lower()

    if TIENE_REGEX:
        tokens = regex_unicode.findall(
            r"[\p{L}\p{M}\p{N}ᵃ-ᶻᴬ-ᵂ⁰-⁹]+(?:[-'][\p{L}\p{M}\p{N}ᵃ-ᶻᴬ-ᵂ⁰-⁹]+)*",
            texto
        )
    else:
        tokens = re.findall(
            r"[A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9ᵃ-ᶻᴬ-ᵂ⁰-⁹]+(?:[-'][A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9ᵃ-ᶻᴬ-ᵂ⁰-⁹]+)*",
            texto
        )

    tokens = [t for t in tokens if t.strip()]

    if quitar_stopwords:
        tokens = [t for t in tokens if t not in STOPWORDS_ES and len(t) > 1]

    return tokens


def metricas_texto(texto: str) -> dict:
    tokens = tokenizar(texto)
    frec = Counter(tokens)
    n_tokens = len(tokens)
    n_unicos = len(frec)
    hapax = sum(1 for _, v in frec.items() if v == 1)

    return {
        "caracteres": len(texto),
        "tokens": n_tokens,
        "palabras_unicas": n_unicos,
        "hapax": hapax,
        "ttr_riqueza_lexica": round(n_unicos / n_tokens, 6) if n_tokens else 0,
        "porcentaje_hapax": round(hapax / n_unicos * 100, 4) if n_unicos else 0,
        "longitud_media_token": round(sum(len(t) for t in tokens) / n_tokens, 4) if n_tokens else 0,
    }


def ratio_seguro(a: float, b: float) -> float:
    return round(a / b, 6) if b else 0


def similitud_texto(a: str, b: str) -> float:
    return round(difflib.SequenceMatcher(None, a, b).ratio(), 6)


def similitud_tokens(tokens_a: list[str], tokens_b: list[str]) -> float:
    return round(difflib.SequenceMatcher(None, tokens_a, tokens_b).ratio(), 6)


def jaccard(set_a: set, set_b: set) -> float:
    union = len(set_a | set_b)
    return round(len(set_a & set_b) / union, 6) if union else 0


def similitud_coseno_counter(c1: Counter, c2: Counter) -> float:
    vocab = set(c1) | set(c2)

    if not vocab:
        return 0

    dot = sum(c1.get(t, 0) * c2.get(t, 0) for t in vocab)
    n1 = math.sqrt(sum(v * v for v in c1.values()))
    n2 = math.sqrt(sum(v * v for v in c2.values()))

    if not n1 or not n2:
        return 0

    return round(dot / (n1 * n2), 6)


# ============================================================
# CARGA
# ============================================================

def cargar_corpus_moderno(carpeta_moderna: Path) -> pd.DataFrame:
    registros = []

    for p in sorted(carpeta_moderna.glob("*_moderno.txt")):
        meta = parse_moderno(p)
        if not meta:
            continue

        registros.append({
            "key": meta["key"],
            "carta_id": meta["carta_id"],
            "nro_carta": meta["nro_carta"],
            "pagina": meta["pagina"],
            "path_moderno": str(p),
            "texto_moderno": leer_texto(p),
        })

    df = pd.DataFrame(registros)

    if df.empty:
        raise FileNotFoundError(f"No se encontraron archivos *_moderno.txt en {carpeta_moderna}")

    return df.sort_values(["nro_carta", "carta_id", "pagina"], na_position="last")


def cargar_corpus_comparacion(carpeta_original: Path, carpeta_moderna: Path) -> pd.DataFrame:
    originales = {}
    modernos = {}

    for p in sorted(carpeta_original.glob("*_page_*.txt")):
        meta = parse_original(p)
        if meta:
            originales[meta["key"]] = meta

    for p in sorted(carpeta_moderna.glob("*_moderno.txt")):
        meta = parse_moderno(p)
        if meta:
            modernos[meta["key"]] = meta

    keys = sorted(set(originales) | set(modernos))
    registros = []

    for key in keys:
        o = originales.get(key, {})
        m = modernos.get(key, {})

        carta_id = o.get("carta_id") or m.get("carta_id")
        pagina = o.get("pagina") or m.get("pagina")
        nro_carta = o.get("nro_carta") or m.get("nro_carta")

        path_original = o.get("path_original")
        path_moderno = m.get("path_moderno")

        registros.append({
            "key": key,
            "carta_id": carta_id,
            "nro_carta": nro_carta,
            "pagina": pagina,
            "path_original": str(path_original) if path_original else "",
            "path_moderno": str(path_moderno) if path_moderno else "",
            "texto_original": leer_texto(path_original) if path_original else "",
            "texto_moderno": leer_texto(path_moderno) if path_moderno else "",
            "existe_original": bool(path_original),
            "existe_moderno": bool(path_moderno),
        })

    df = pd.DataFrame(registros)

    if df.empty:
        raise FileNotFoundError("No se pudo construir corpus comparativo.")

    return df.sort_values(["nro_carta", "carta_id", "pagina"], na_position="last")


# ============================================================
# ANÁLISIS MODERNO
# ============================================================

def moderno_resumen_global(df_mod: pd.DataFrame, carpeta_moderna: Path) -> pd.DataFrame:
    texto = "\n\n".join(df_mod["texto_moderno"].fillna("").tolist())
    mt = metricas_texto(texto)

    return pd.DataFrame([{
        "carpeta_moderna": str(carpeta_moderna),
        "total_cartas": df_mod["carta_id"].nunique(),
        "total_paginas": len(df_mod),
        **mt,
    }])


def moderno_resumen_por_pagina(df_mod: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for _, r in df_mod.iterrows():
        mt = metricas_texto(r["texto_moderno"])
        filas.append({
            "key": r["key"],
            "carta_id": r["carta_id"],
            "nro_carta": r["nro_carta"],
            "pagina": r["pagina"],
            **mt,
            "path_moderno": r["path_moderno"],
        })

    return pd.DataFrame(filas)


def moderno_resumen_por_carta(df_mod: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for carta_id, g in df_mod.groupby("carta_id", sort=False):
        texto = "\n\n".join(g["texto_moderno"].fillna("").tolist())
        mt = metricas_texto(texto)
        filas.append({
            "carta_id": carta_id,
            "nro_carta": g["nro_carta"].iloc[0],
            "total_paginas": len(g),
            **mt,
        })

    return pd.DataFrame(filas)


def moderno_frecuencias_por_pagina(df_mod: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for _, r in df_mod.iterrows():
        tokens = tokenizar(r["texto_moderno"])
        total = len(tokens)
        frec = Counter(tokens)

        for rank, (palabra, frecuencia) in enumerate(frec.most_common(TOP_FRECUENCIAS_PAGINA), start=1):
            filas.append({
                "key": r["key"],
                "carta_id": r["carta_id"],
                "nro_carta": r["nro_carta"],
                "pagina": r["pagina"],
                "rank": rank,
                "palabra": palabra,
                "frecuencia": frecuencia,
                "porcentaje": round(frecuencia / total * 100, 6) if total else 0,
            })

    return pd.DataFrame(filas)


def moderno_frecuencias_por_carta(df_mod: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for carta_id, g in df_mod.groupby("carta_id", sort=False):
        texto = "\n\n".join(g["texto_moderno"].fillna("").tolist())
        tokens = tokenizar(texto)
        total = len(tokens)
        frec = Counter(tokens)

        for rank, (palabra, frecuencia) in enumerate(frec.most_common(TOP_FRECUENCIAS_CARTA), start=1):
            filas.append({
                "carta_id": carta_id,
                "nro_carta": g["nro_carta"].iloc[0],
                "rank": rank,
                "palabra": palabra,
                "frecuencia": frecuencia,
                "porcentaje": round(frecuencia / total * 100, 6) if total else 0,
            })

    return pd.DataFrame(filas)


def moderno_frecuencia_global(df_mod: pd.DataFrame) -> pd.DataFrame:
    texto = "\n\n".join(df_mod["texto_moderno"].fillna("").tolist())
    tokens = tokenizar(texto)
    total = len(tokens)
    frec = Counter(tokens)

    return pd.DataFrame([
        {
            "rank": rank,
            "palabra": palabra,
            "frecuencia": frecuencia,
            "porcentaje": round(frecuencia / total * 100, 6) if total else 0,
        }
        for rank, (palabra, frecuencia) in enumerate(frec.most_common(), start=1)
    ])


def moderno_tfidf_por_carta(df_mod: pd.DataFrame) -> pd.DataFrame:
    if not TIENE_SKLEARN:
        return pd.DataFrame()

    docs = []
    carta_ids = []
    nro_cartas = []

    for carta_id, g in df_mod.groupby("carta_id", sort=False):
        texto = "\n\n".join(g["texto_moderno"].fillna("").tolist())
        docs.append(" ".join(tokenizar(texto)))
        carta_ids.append(carta_id)
        nro_cartas.append(g["nro_carta"].iloc[0])

    if not docs:
        return pd.DataFrame()

    vectorizer = TfidfVectorizer(
        lowercase=False,
        token_pattern=r"(?u)\b\w+\b",
        min_df=1
    )

    X = vectorizer.fit_transform(docs)
    terms = vectorizer.get_feature_names_out()

    filas = []

    for i, carta_id in enumerate(carta_ids):
        row = X[i].toarray().ravel()
        top_idx = row.argsort()[::-1][:TOP_TFIDF_CARTA]

        for rank, idx in enumerate(top_idx, start=1):
            score = row[idx]
            if score <= 0:
                continue
            filas.append({
                "carta_id": carta_id,
                "nro_carta": nro_cartas[i],
                "rank": rank,
                "termino": terms[idx],
                "tfidf": round(float(score), 8),
            })

    return pd.DataFrame(filas)


def moderno_asociaciones_por_carta(df_mod: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for carta_id, g in df_mod.groupby("carta_id", sort=False):
        texto = "\n\n".join(g["texto_moderno"].fillna("").tolist())
        tokens = tokenizar(texto, quitar_stopwords=QUITAR_STOPWORDS_EN_REDES)
        cooc = Counter()

        for i, token in enumerate(tokens):
            ventana = tokens[i + 1:i + 1 + VENTANA_ASOCIACION]
            for otro in ventana:
                if token == otro:
                    continue
                a, b = sorted([token, otro])
                cooc[(a, b)] += 1

        for rank, ((a, b), peso) in enumerate(cooc.most_common(TOP_ASOCIACIONES_CARTA), start=1):
            filas.append({
                "carta_id": carta_id,
                "nro_carta": g["nro_carta"].iloc[0],
                "rank": rank,
                "palabra_1": a,
                "palabra_2": b,
                "peso_coocurrencia": peso,
                "ventana": VENTANA_ASOCIACION,
            })

    return pd.DataFrame(filas)


def moderno_asociaciones_global(df_mod: pd.DataFrame) -> pd.DataFrame:
    texto = "\n\n".join(df_mod["texto_moderno"].fillna("").tolist())
    tokens = tokenizar(texto, quitar_stopwords=QUITAR_STOPWORDS_EN_REDES)
    cooc = Counter()

    for i, token in enumerate(tokens):
        ventana = tokens[i + 1:i + 1 + VENTANA_ASOCIACION]
        for otro in ventana:
            if token == otro:
                continue
            a, b = sorted([token, otro])
            cooc[(a, b)] += 1

    return pd.DataFrame([
        {
            "rank": rank,
            "palabra_1": a,
            "palabra_2": b,
            "peso_coocurrencia": peso,
            "ventana": VENTANA_ASOCIACION,
        }
        for rank, ((a, b), peso) in enumerate(cooc.most_common(), start=1)
    ])


def moderno_longitud_oraciones(df_mod: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for _, r in df_mod.iterrows():
        texto = r["texto_moderno"]
        oraciones = [o.strip() for o in re.split(r"[.!?]+", texto) if o.strip()]
        longitudes = [len(tokenizar(o)) for o in oraciones if tokenizar(o)]

        filas.append({
            "key": r["key"],
            "carta_id": r["carta_id"],
            "nro_carta": r["nro_carta"],
            "pagina": r["pagina"],
            "n_oraciones": len(longitudes),
            "promedio_tokens_oracion": round(sum(longitudes) / len(longitudes), 4) if longitudes else 0,
            "max_tokens_oracion": max(longitudes) if longitudes else 0,
            "min_tokens_oracion": min(longitudes) if longitudes else 0,
        })

    return pd.DataFrame(filas)


# ============================================================
# REEMPLAZOS
# ============================================================

def limpiar_token_para_reemplazo(t: str) -> str:
    t = str(t).strip().lower()
    t = t.strip(".,;:!?¡¿()[]{}\"'“”‘’")
    return t


def extraer_reemplazos_pagina(tokens_original: list[str], tokens_moderno: list[str]) -> list[dict]:
    sm = difflib.SequenceMatcher(None, tokens_original, tokens_moderno)
    reemplazos = []

    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        orig_span = tokens_original[i1:i2]
        mod_span = tokens_moderno[j1:j2]

        if tag == "equal":
            continue

        orig_bloque = " ".join(orig_span)
        mod_bloque = " ".join(mod_span)

        if tag == "replace":
            max_len = max(len(orig_span), len(mod_span))

            for k in range(max_len):
                palabra_original = orig_span[k] if k < len(orig_span) else ""
                palabra_moderna = mod_span[k] if k < len(mod_span) else ""

                if palabra_original == palabra_moderna:
                    continue

                reemplazos.append({
                    "tipo_cambio": "replace_token",
                    "palabra_original": palabra_original,
                    "palabra_moderna": palabra_moderna,
                    "bloque_original": orig_bloque,
                    "bloque_moderno": mod_bloque,
                    "n_tokens_original_bloque": len(orig_span),
                    "n_tokens_moderno_bloque": len(mod_span),
                })

            reemplazos.append({
                "tipo_cambio": "replace_bloque",
                "palabra_original": orig_bloque,
                "palabra_moderna": mod_bloque,
                "bloque_original": orig_bloque,
                "bloque_moderno": mod_bloque,
                "n_tokens_original_bloque": len(orig_span),
                "n_tokens_moderno_bloque": len(mod_span),
            })

        elif tag == "delete":
            for palabra_original in orig_span:
                reemplazos.append({
                    "tipo_cambio": "delete_token",
                    "palabra_original": palabra_original,
                    "palabra_moderna": "",
                    "bloque_original": orig_bloque,
                    "bloque_moderno": "",
                    "n_tokens_original_bloque": len(orig_span),
                    "n_tokens_moderno_bloque": 0,
                })

        elif tag == "insert":
            for palabra_moderna in mod_span:
                reemplazos.append({
                    "tipo_cambio": "insert_token",
                    "palabra_original": "",
                    "palabra_moderna": palabra_moderna,
                    "bloque_original": "",
                    "bloque_moderno": mod_bloque,
                    "n_tokens_original_bloque": 0,
                    "n_tokens_moderno_bloque": len(mod_span),
                })

    return reemplazos


def analizar_reemplazos(df_comp: pd.DataFrame) -> dict[str, pd.DataFrame]:
    detalle = []

    for _, r in df_comp.iterrows():
        tok_o = tokenizar(r["texto_original"], quitar_stopwords=False)
        tok_m = tokenizar(r["texto_moderno"], quitar_stopwords=False)
        cambios = extraer_reemplazos_pagina(tok_o, tok_m)

        for c in cambios:
            po = limpiar_token_para_reemplazo(c["palabra_original"])
            pm = limpiar_token_para_reemplazo(c["palabra_moderna"])

            detalle.append({
                "key": r["key"],
                "carta_id": r["carta_id"],
                "nro_carta": r["nro_carta"],
                "pagina": r["pagina"],
                "tipo_cambio": c["tipo_cambio"],
                "palabra_original": po,
                "palabra_moderna": pm,
                "bloque_original": c["bloque_original"],
                "bloque_moderno": c["bloque_moderno"],
                "n_tokens_original_bloque": c["n_tokens_original_bloque"],
                "n_tokens_moderno_bloque": c["n_tokens_moderno_bloque"],
            })

    df_detalle = pd.DataFrame(detalle)

    if df_detalle.empty:
        vacio = pd.DataFrame()
        return {
            "reemplazos_detalle": vacio,
            "reemplazos_global": vacio,
            "reemplazos_por_carta": vacio,
            "reemplazos_por_pagina": vacio,
            "reemplazos_resumen_carta": vacio,
            "reemplazos_resumen_pagina": vacio,
        }

    df_token = df_detalle[df_detalle["tipo_cambio"].isin(["replace_token", "delete_token", "insert_token"])].copy()

    reemplazos_global = (
        df_token
        .groupby(["tipo_cambio", "palabra_original", "palabra_moderna"], as_index=False)
        .size()
        .rename(columns={"size": "frecuencia"})
        .sort_values("frecuencia", ascending=False)
        .head(TOP_REEMPLAZOS_GLOBAL)
    )

    reemplazos_por_carta = (
        df_token
        .groupby(["carta_id", "nro_carta", "tipo_cambio", "palabra_original", "palabra_moderna"], as_index=False)
        .size()
        .rename(columns={"size": "frecuencia"})
        .sort_values(["nro_carta", "frecuencia"], ascending=[True, False])
        .groupby("carta_id", as_index=False)
        .head(TOP_REEMPLAZOS_CARTA)
    )

    reemplazos_por_pagina = (
        df_token
        .groupby(["key", "carta_id", "nro_carta", "pagina", "tipo_cambio", "palabra_original", "palabra_moderna"], as_index=False)
        .size()
        .rename(columns={"size": "frecuencia"})
        .sort_values(["nro_carta", "pagina", "frecuencia"], ascending=[True, True, False])
        .groupby("key", as_index=False)
        .head(TOP_REEMPLAZOS_PAGINA)
    )

    resumen_carta = (
        df_token
        .groupby(["carta_id", "nro_carta", "tipo_cambio"], as_index=False)
        .size()
        .rename(columns={"size": "total_cambios"})
        .pivot_table(index=["carta_id", "nro_carta"], columns="tipo_cambio", values="total_cambios", fill_value=0)
        .reset_index()
    )

    resumen_pagina = (
        df_token
        .groupby(["key", "carta_id", "nro_carta", "pagina", "tipo_cambio"], as_index=False)
        .size()
        .rename(columns={"size": "total_cambios"})
        .pivot_table(index=["key", "carta_id", "nro_carta", "pagina"], columns="tipo_cambio", values="total_cambios", fill_value=0)
        .reset_index()
    )

    return {
        "reemplazos_detalle": df_detalle,
        "reemplazos_global": reemplazos_global,
        "reemplazos_por_carta": reemplazos_por_carta,
        "reemplazos_por_pagina": reemplazos_por_pagina,
        "reemplazos_resumen_carta": resumen_carta,
        "reemplazos_resumen_pagina": resumen_pagina,
    }


# ============================================================
# COMPARACIÓN
# ============================================================

def comparacion_por_pagina(df: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for _, r in df.iterrows():
        original = r["texto_original"]
        moderno = r["texto_moderno"]

        mt_o = metricas_texto(original)
        mt_m = metricas_texto(moderno)

        tok_o = tokenizar(original)
        tok_m = tokenizar(moderno)

        c_o = Counter(tok_o)
        c_m = Counter(tok_m)

        vocab_o = set(c_o)
        vocab_m = set(c_m)

        filas.append({
            "key": r["key"],
            "carta_id": r["carta_id"],
            "nro_carta": r["nro_carta"],
            "pagina": r["pagina"],
            "existe_original": r["existe_original"],
            "existe_moderno": r["existe_moderno"],
            "original_caracteres": mt_o["caracteres"],
            "original_tokens": mt_o["tokens"],
            "original_palabras_unicas": mt_o["palabras_unicas"],
            "original_hapax": mt_o["hapax"],
            "original_ttr": mt_o["ttr_riqueza_lexica"],
            "moderno_caracteres": mt_m["caracteres"],
            "moderno_tokens": mt_m["tokens"],
            "moderno_palabras_unicas": mt_m["palabras_unicas"],
            "moderno_hapax": mt_m["hapax"],
            "moderno_ttr": mt_m["ttr_riqueza_lexica"],
            "dif_caracteres_mod_menos_orig": mt_m["caracteres"] - mt_o["caracteres"],
            "dif_tokens_mod_menos_orig": mt_m["tokens"] - mt_o["tokens"],
            "dif_palabras_unicas_mod_menos_orig": mt_m["palabras_unicas"] - mt_o["palabras_unicas"],
            "ratio_tokens_mod_orig": ratio_seguro(mt_m["tokens"], mt_o["tokens"]),
            "ratio_caracteres_mod_orig": ratio_seguro(mt_m["caracteres"], mt_o["caracteres"]),
            "similitud_caracteres_sequence": similitud_texto(original, moderno),
            "similitud_tokens_sequence": similitud_tokens(tok_o, tok_m),
            "jaccard_vocabulario": jaccard(vocab_o, vocab_m),
            "coseno_frecuencias": similitud_coseno_counter(c_o, c_m),
            "vocab_comun": len(vocab_o & vocab_m),
            "vocab_solo_original": len(vocab_o - vocab_m),
            "vocab_solo_moderno": len(vocab_m - vocab_o),
        })

    return pd.DataFrame(filas)


def comparacion_por_carta(df: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for carta_id, g in df.groupby("carta_id", sort=False):
        texto_o = "\n\n".join(g["texto_original"].fillna("").tolist())
        texto_m = "\n\n".join(g["texto_moderno"].fillna("").tolist())

        mt_o = metricas_texto(texto_o)
        mt_m = metricas_texto(texto_m)

        tok_o = tokenizar(texto_o)
        tok_m = tokenizar(texto_m)

        c_o = Counter(tok_o)
        c_m = Counter(tok_m)

        vocab_o = set(c_o)
        vocab_m = set(c_m)

        filas.append({
            "carta_id": carta_id,
            "nro_carta": g["nro_carta"].iloc[0],
            "total_paginas": len(g),
            "original_caracteres": mt_o["caracteres"],
            "original_tokens": mt_o["tokens"],
            "original_palabras_unicas": mt_o["palabras_unicas"],
            "original_hapax": mt_o["hapax"],
            "original_ttr": mt_o["ttr_riqueza_lexica"],
            "moderno_caracteres": mt_m["caracteres"],
            "moderno_tokens": mt_m["tokens"],
            "moderno_palabras_unicas": mt_m["palabras_unicas"],
            "moderno_hapax": mt_m["hapax"],
            "moderno_ttr": mt_m["ttr_riqueza_lexica"],
            "dif_caracteres_mod_menos_orig": mt_m["caracteres"] - mt_o["caracteres"],
            "dif_tokens_mod_menos_orig": mt_m["tokens"] - mt_o["tokens"],
            "dif_palabras_unicas_mod_menos_orig": mt_m["palabras_unicas"] - mt_o["palabras_unicas"],
            "ratio_tokens_mod_orig": ratio_seguro(mt_m["tokens"], mt_o["tokens"]),
            "ratio_caracteres_mod_orig": ratio_seguro(mt_m["caracteres"], mt_o["caracteres"]),
            "similitud_caracteres_sequence": similitud_texto(texto_o, texto_m),
            "similitud_tokens_sequence": similitud_tokens(tok_o, tok_m),
            "jaccard_vocabulario": jaccard(vocab_o, vocab_m),
            "coseno_frecuencias": similitud_coseno_counter(c_o, c_m),
            "vocab_comun": len(vocab_o & vocab_m),
            "vocab_solo_original": len(vocab_o - vocab_m),
            "vocab_solo_moderno": len(vocab_m - vocab_o),
        })

    return pd.DataFrame(filas)


def comparacion_global(df: pd.DataFrame, carpeta_original: Path, carpeta_moderna: Path) -> pd.DataFrame:
    texto_o = "\n\n".join(df["texto_original"].fillna("").tolist())
    texto_m = "\n\n".join(df["texto_moderno"].fillna("").tolist())

    mt_o = metricas_texto(texto_o)
    mt_m = metricas_texto(texto_m)

    tok_o = tokenizar(texto_o)
    tok_m = tokenizar(texto_m)

    c_o = Counter(tok_o)
    c_m = Counter(tok_m)

    vocab_o = set(c_o)
    vocab_m = set(c_m)

    return pd.DataFrame([{
        "carpeta_original": str(carpeta_original),
        "carpeta_moderna": str(carpeta_moderna),
        "total_paginas": len(df),
        "total_cartas": df["carta_id"].nunique(),
        "paginas_con_original": int(df["existe_original"].sum()),
        "paginas_con_moderno": int(df["existe_moderno"].sum()),
        "original_caracteres": mt_o["caracteres"],
        "original_tokens": mt_o["tokens"],
        "original_palabras_unicas": mt_o["palabras_unicas"],
        "original_hapax": mt_o["hapax"],
        "original_ttr": mt_o["ttr_riqueza_lexica"],
        "moderno_caracteres": mt_m["caracteres"],
        "moderno_tokens": mt_m["tokens"],
        "moderno_palabras_unicas": mt_m["palabras_unicas"],
        "moderno_hapax": mt_m["hapax"],
        "moderno_ttr": mt_m["ttr_riqueza_lexica"],
        "dif_caracteres_mod_menos_orig": mt_m["caracteres"] - mt_o["caracteres"],
        "dif_tokens_mod_menos_orig": mt_m["tokens"] - mt_o["tokens"],
        "dif_palabras_unicas_mod_menos_orig": mt_m["palabras_unicas"] - mt_o["palabras_unicas"],
        "ratio_tokens_mod_orig": ratio_seguro(mt_m["tokens"], mt_o["tokens"]),
        "ratio_caracteres_mod_orig": ratio_seguro(mt_m["caracteres"], mt_o["caracteres"]),
        "similitud_caracteres_total": similitud_texto(texto_o, texto_m),
        "similitud_tokens_total": similitud_tokens(tok_o, tok_m),
        "jaccard_vocabulario_total": jaccard(vocab_o, vocab_m),
        "coseno_frecuencias_total": similitud_coseno_counter(c_o, c_m),
        "vocab_comun_total": len(vocab_o & vocab_m),
        "vocab_solo_original_total": len(vocab_o - vocab_m),
        "vocab_solo_moderno_total": len(vocab_m - vocab_o),
    }])


def cambios_frecuencia_por_carta(df: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for carta_id, g in df.groupby("carta_id", sort=False):
        texto_o = "\n\n".join(g["texto_original"].fillna("").tolist())
        texto_m = "\n\n".join(g["texto_moderno"].fillna("").tolist())

        c_o = Counter(tokenizar(texto_o))
        c_m = Counter(tokenizar(texto_m))

        vocab = set(c_o) | set(c_m)

        tmp = []

        for palabra in vocab:
            fo = c_o.get(palabra, 0)
            fm = c_m.get(palabra, 0)

            tmp.append({
                "carta_id": carta_id,
                "nro_carta": g["nro_carta"].iloc[0],
                "palabra": palabra,
                "frecuencia_original": fo,
                "frecuencia_moderna": fm,
                "diferencia_mod_menos_orig": fm - fo,
                "abs_diferencia": abs(fm - fo),
                "ratio_mod_orig_suavizado": round((fm + 1) / (fo + 1), 6),
            })

        tmp = sorted(tmp, key=lambda x: (-x["abs_diferencia"], x["palabra"]))[:TOP_CAMBIOS_CARTA]
        filas.extend(tmp)

    return pd.DataFrame(filas)


def vocabulario_exclusivo(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    texto_o = "\n\n".join(df["texto_original"].fillna("").tolist())
    texto_m = "\n\n".join(df["texto_moderno"].fillna("").tolist())

    c_o = Counter(tokenizar(texto_o))
    c_m = Counter(tokenizar(texto_m))

    solo_o = set(c_o) - set(c_m)
    solo_m = set(c_m) - set(c_o)

    df_solo_o = pd.DataFrame([
        {"palabra": p, "frecuencia_original": c_o[p]}
        for p in solo_o
    ]).sort_values("frecuencia_original", ascending=False).head(TOP_VOCAB_EXCLUSIVO)

    df_solo_m = pd.DataFrame([
        {"palabra": p, "frecuencia_moderna": c_m[p]}
        for p in solo_m
    ]).sort_values("frecuencia_moderna", ascending=False).head(TOP_VOCAB_EXCLUSIVO)

    return df_solo_o, df_solo_m


# ============================================================
# HTML
# ============================================================

def generar_html_comparaciones(df: pd.DataFrame) -> None:
    for _, r in df.iterrows():
        carta_id = r["carta_id"]
        pagina = int(r["pagina"])
        base = limpiar_nombre_archivo(f"{carta_id}_page_{pagina:03d}")

        original = r["texto_original"] or ""
        moderno = r["texto_moderno"] or ""

        doc = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Comparación - {html.escape(carta_id)} - Página {pagina}</title>
<style>
body {{
    font-family: Arial, sans-serif;
    margin: 24px;
}}
.container {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 24px;
}}
.panel {{
    border: 1px solid #ddd;
    padding: 16px;
    border-radius: 8px;
}}
.texto {{
    font-family: Georgia, "Times New Roman", serif;
    font-size: 18px;
    line-height: 1.55;
    white-space: pre-wrap;
}}
</style>
</head>
<body>
<h1>Comparación original vs nueva versión</h1>
<p><strong>Carta:</strong> {html.escape(str(carta_id))} | <strong>Página:</strong> {pagina}</p>
<div class="container">
<div class="panel">
<h2>Original diplomático</h2>
<div class="texto">{html.escape(original)}</div>
</div>
<div class="panel">
<h2>Versión moderna / normalizada</h2>
<div class="texto">{html.escape(moderno)}</div>
</div>
</div>
</body>
</html>
"""

        salida = CARPETA_COMP_HTML / f"{base}_comparacion.html"
        salida.write_text(doc, encoding="utf-8")


# ============================================================
# GRÁFICOS Y REDES
# ============================================================

def grafico_barras(df: pd.DataFrame, x: str, y: str, titulo: str, salida: Path, top_n: int = 30) -> None:
    if df.empty:
        return

    d = df.head(top_n).copy()

    plt.figure(figsize=(12, 6))
    plt.bar(d[x].astype(str), d[y])
    plt.title(titulo)
    plt.xticks(rotation=75, ha="right")
    plt.tight_layout()
    plt.savefig(salida, dpi=160)
    plt.close()


def graficos_moderno(resumen_carta: pd.DataFrame, frec_global: pd.DataFrame) -> None:
    if not resumen_carta.empty:
        d = resumen_carta.sort_values("nro_carta")

        plt.figure(figsize=(12, 6))
        plt.bar(d["nro_carta"].astype(str), d["tokens"])
        plt.title("Tokens por carta - versión moderna")
        plt.xlabel("Carta")
        plt.ylabel("Tokens")
        plt.xticks(rotation=75)
        plt.tight_layout()
        plt.savefig(CARPETA_MODERNO_GRAFICOS / "tokens_por_carta_moderno.png", dpi=160)
        plt.close()

        plt.figure(figsize=(12, 6))
        plt.bar(d["nro_carta"].astype(str), d["ttr_riqueza_lexica"])
        plt.title("Riqueza léxica TTR por carta - versión moderna")
        plt.xlabel("Carta")
        plt.ylabel("TTR")
        plt.xticks(rotation=75)
        plt.tight_layout()
        plt.savefig(CARPETA_MODERNO_GRAFICOS / "ttr_por_carta_moderno.png", dpi=160)
        plt.close()

        plt.figure(figsize=(12, 6))
        plt.bar(d["nro_carta"].astype(str), d["hapax"])
        plt.title("Hapax por carta - versión moderna")
        plt.xlabel("Carta")
        plt.ylabel("Hapax")
        plt.xticks(rotation=75)
        plt.tight_layout()
        plt.savefig(CARPETA_MODERNO_GRAFICOS / "hapax_por_carta_moderno.png", dpi=160)
        plt.close()

    if not frec_global.empty:
        grafico_barras(
            frec_global,
            "palabra",
            "frecuencia",
            "Top 30 palabras de la versión moderna",
            CARPETA_MODERNO_GRAFICOS / "top_30_palabras_moderno.png",
            top_n=30
        )


def graficos_reemplazos(reemplazos_global: pd.DataFrame) -> None:
    if reemplazos_global.empty:
        return

    d = reemplazos_global[reemplazos_global["tipo_cambio"] == "replace_token"].copy().head(30)

    if d.empty:
        return

    d["par"] = d["palabra_original"].astype(str) + " → " + d["palabra_moderna"].astype(str)

    grafico_barras(
        d,
        "par",
        "frecuencia",
        "Top 30 reemplazos original → moderno",
        CARPETA_REEMPLAZOS / "top_30_reemplazos.png",
        top_n=30
    )


def graficos_comparacion(resumen_carta: pd.DataFrame, cambios: pd.DataFrame) -> None:
    if not resumen_carta.empty:
        d = resumen_carta.sort_values("nro_carta")

        plt.figure(figsize=(12, 6))
        plt.plot(d["nro_carta"].astype(str), d["original_tokens"], marker="o", label="Original")
        plt.plot(d["nro_carta"].astype(str), d["moderno_tokens"], marker="o", label="Moderno")
        plt.title("Tokens por carta: original vs moderno")
        plt.xlabel("Carta")
        plt.ylabel("Tokens")
        plt.legend()
        plt.xticks(rotation=75)
        plt.tight_layout()
        plt.savefig(CARPETA_COMP_GRAFICOS / "tokens_original_vs_moderno_por_carta.png", dpi=160)
        plt.close()

        plt.figure(figsize=(12, 6))
        plt.bar(d["nro_carta"].astype(str), d["similitud_tokens_sequence"])
        plt.title("Similitud de tokens entre original y moderno por carta")
        plt.xlabel("Carta")
        plt.ylabel("Similitud")
        plt.xticks(rotation=75)
        plt.tight_layout()
        plt.savefig(CARPETA_COMP_GRAFICOS / "similitud_tokens_por_carta.png", dpi=160)
        plt.close()

    if not cambios.empty:
        total_cambios = (
            cambios
            .groupby("palabra", as_index=False)["abs_diferencia"]
            .sum()
            .sort_values("abs_diferencia", ascending=False)
            .head(30)
        )

        grafico_barras(
            total_cambios,
            "palabra",
            "abs_diferencia",
            "Top 30 palabras con mayor cambio absoluto",
            CARPETA_COMP_GRAFICOS / "top_30_cambios_frecuencia.png",
            top_n=30
        )


def generar_nubes_moderno(df_mod: pd.DataFrame) -> None:
    if not GENERAR_NUBES or not TIENE_WORDCLOUD:
        return

    cartas = list(df_mod.groupby("carta_id", sort=False))[:TOP_NUBES_CARTAS]

    for carta_id, g in cartas:
        texto = "\n\n".join(g["texto_moderno"].fillna("").tolist())
        tokens = tokenizar(texto)
        frec = Counter(tokens)

        if not frec:
            continue

        wc = WordCloud(
            width=1400,
            height=800,
            background_color="white",
            collocations=False
        ).generate_from_frequencies(frec)

        salida = CARPETA_MODERNO_NUBES / f"{limpiar_nombre_archivo(carta_id)}_wordcloud_moderno.png"
        wc.to_file(str(salida))


def dibujar_red_asociaciones(df_edges: pd.DataFrame, salida: Path, titulo: str,
                             top_edges: int, top_nodes: int) -> None:
    if not TIENE_NETWORKX or df_edges.empty:
        return

    d = df_edges.copy()
    d = d[d["peso_coocurrencia"] >= MIN_PESO_ARISTA]

    if d.empty:
        return

    d = d.sort_values("peso_coocurrencia", ascending=False).head(top_edges)

    G = nx.Graph()

    for _, r in d.iterrows():
        a = str(r["palabra_1"])
        b = str(r["palabra_2"])
        w = float(r["peso_coocurrencia"])
        G.add_edge(a, b, weight=w)

    if G.number_of_nodes() > top_nodes:
        grados = dict(G.degree(weight="weight"))
        nodos_top = sorted(grados, key=grados.get, reverse=True)[:top_nodes]
        G = G.subgraph(nodos_top).copy()

    if G.number_of_edges() == 0:
        return

    plt.figure(figsize=(14, 10))
    pos = nx.spring_layout(G, k=0.7, seed=42, weight="weight")

    weights = [G[u][v]["weight"] for u, v in G.edges()]
    max_w = max(weights) if weights else 1
    widths = [0.5 + 4 * (w / max_w) for w in weights]

    grados = dict(G.degree(weight="weight"))
    max_g = max(grados.values()) if grados else 1
    sizes = [300 + 1700 * (grados[n] / max_g) for n in G.nodes()]

    nx.draw_networkx_nodes(G, pos, node_size=sizes, alpha=0.85)
    nx.draw_networkx_edges(G, pos, width=widths, alpha=0.35)
    nx.draw_networkx_labels(G, pos, font_size=9)

    plt.title(titulo)
    plt.axis("off")
    plt.tight_layout()
    plt.savefig(salida, dpi=180)
    plt.close()


def generar_redes_asociaciones(asoc_carta: pd.DataFrame, asoc_global: pd.DataFrame) -> None:
    if not GENERAR_REDES_ASOCIACIONES or not TIENE_NETWORKX:
        return

    dibujar_red_asociaciones(
        asoc_global,
        CARPETA_MODERNO_REDES / "red_asociaciones_global_moderno.png",
        "Red global de asociaciones - versión moderna",
        TOP_ARISTAS_RED_GLOBAL,
        TOP_NODOS_RED_GLOBAL
    )

    if asoc_carta.empty:
        return

    for carta_id, g in asoc_carta.groupby("carta_id", sort=False):
        nro = g["nro_carta"].iloc[0]
        salida = CARPETA_MODERNO_REDES / f"{int(nro):02d}_{limpiar_nombre_archivo(carta_id, 80)}_red_asociaciones.png"

        dibujar_red_asociaciones(
            g,
            salida,
            f"Red de asociaciones - Carta {nro}",
            TOP_ARISTAS_RED_CARTA,
            TOP_NODOS_RED_CARTA
        )


def generar_red_reemplazos(reemplazos_global: pd.DataFrame) -> None:
    if not GENERAR_RED_REEMPLAZOS or not TIENE_NETWORKX or reemplazos_global.empty:
        return

    d = reemplazos_global.copy()
    d = d[d["tipo_cambio"] == "replace_token"]
    d = d[(d["palabra_original"] != "") & (d["palabra_moderna"] != "")]
    d = d.sort_values("frecuencia", ascending=False).head(TOP_ARISTAS_RED_REEMPLAZOS)

    if d.empty:
        return

    G = nx.DiGraph()

    for _, r in d.iterrows():
        a = "O: " + str(r["palabra_original"])
        b = "M: " + str(r["palabra_moderna"])
        w = float(r["frecuencia"])
        G.add_edge(a, b, weight=w)

    plt.figure(figsize=(16, 12))
    pos = nx.spring_layout(G, k=0.85, seed=42, weight="weight")

    weights = [G[u][v]["weight"] for u, v in G.edges()]
    max_w = max(weights) if weights else 1
    widths = [0.5 + 5 * (w / max_w) for w in weights]

    grados = dict(G.degree(weight="weight"))
    max_g = max(grados.values()) if grados else 1
    sizes = [250 + 1500 * (grados[n] / max_g) for n in G.nodes()]

    nx.draw_networkx_nodes(G, pos, node_size=sizes, alpha=0.85)
    nx.draw_networkx_edges(G, pos, width=widths, alpha=0.35, arrows=True, arrowsize=12)
    nx.draw_networkx_labels(G, pos, font_size=8)

    plt.title("Red de reemplazos original → moderno")
    plt.axis("off")
    plt.tight_layout()
    plt.savefig(CARPETA_REEMPLAZOS / "red_reemplazos_original_moderno.png", dpi=180)
    plt.close()


# ============================================================
# EXCEL
# ============================================================

def exportar_excel(path_excel: Path, hojas: dict[str, pd.DataFrame]) -> None:
    path_excel.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path_excel, engine="openpyxl") as writer:
        for nombre, df in hojas.items():
            sheet = nombre[:31]
            if df is None or df.empty:
                pd.DataFrame({"mensaje": ["Sin datos"]}).to_excel(writer, sheet_name=sheet, index=False)
            else:
                df.to_excel(writer, sheet_name=sheet, index=False)

    aplicar_formato_excel(path_excel)


def exportar_excel_df_grande(path_excel: Path, df: pd.DataFrame, base_sheet: str = "detalle") -> None:
    path_excel.parent.mkdir(parents=True, exist_ok=True)
    max_filas = 1_000_000

    with pd.ExcelWriter(path_excel, engine="openpyxl") as writer:
        if df is None or df.empty:
            pd.DataFrame({"mensaje": ["Sin datos"]}).to_excel(writer, sheet_name=base_sheet[:31], index=False)
        else:
            parte = 1
            for inicio in range(0, len(df), max_filas):
                fin = min(inicio + max_filas, len(df))
                sheet_name = f"{base_sheet}_{parte:02d}"[:31]
                df.iloc[inicio:fin].to_excel(writer, sheet_name=sheet_name, index=False)
                parte += 1

    aplicar_formato_excel(path_excel)


def aplicar_formato_excel(path_excel: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = load_workbook(path_excel)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(val), 80))
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 55))

        wb.save(path_excel)

    except Exception as e:
        print(f"No se pudo aplicar formato al Excel {path_excel}: {e}")


def guardar_resumenes(carpeta_original: Path, carpeta_moderna: Path) -> None:
    texto_moderno = f"""
ANÁLISIS SOLO DE LA VERSIÓN MODERNA / NORMALIZADA

Carpeta moderna analizada:
{carpeta_moderna}

Salida:
{CARPETA_ANALISIS_MODERNO}

No se generan CSV.

Incluye:
- Excel de análisis moderno.
- Reemplazos original → moderno en Excel.
- Redes de asociaciones.
- Red de reemplazos.
- Nubes de palabras.
- Gráficos.
""".strip()

    (CARPETA_ANALISIS_MODERNO / "resumen_analisis_moderno.txt").write_text(texto_moderno, encoding="utf-8")

    texto_comp = f"""
COMPARACIÓN ORIGINAL VS MODERNO

Carpeta original:
{carpeta_original}

Carpeta moderna:
{carpeta_moderna}

Salida:
{CARPETA_COMPARACION}

No se generan CSV.
""".strip()

    (CARPETA_COMPARACION / "resumen_metodologico_comparacion.txt").write_text(texto_comp, encoding="utf-8")


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    print("=" * 72)
    print("ANÁLISIS MODERNO + REEMPLAZOS + REDES + COMPARACIÓN")
    print("=" * 72)

    crear_carpetas()

    carpeta_original = detectar_carpeta_original()
    carpeta_moderna = detectar_carpeta_moderna()

    print(f"Carpeta original detectada: {carpeta_original}")
    print(f"Carpeta moderna detectada: {carpeta_moderna}")
    print(f"NetworkX disponible: {TIENE_NETWORKX}")
    print(f"WordCloud disponible: {TIENE_WORDCLOUD}")
    print(f"Scikit-learn disponible: {TIENE_SKLEARN}")

    # A) Análisis moderno.
    print("\n[A] Cargando corpus moderno...")
    df_mod = cargar_corpus_moderno(carpeta_moderna)

    print(f"Páginas modernas: {len(df_mod)}")
    print(f"Cartas modernas: {df_mod['carta_id'].nunique()}")

    print("Calculando análisis solo moderno...")
    mod_global = moderno_resumen_global(df_mod, carpeta_moderna)
    mod_pagina = moderno_resumen_por_pagina(df_mod)
    mod_carta = moderno_resumen_por_carta(df_mod)
    mod_frec_pagina = moderno_frecuencias_por_pagina(df_mod)
    mod_frec_carta = moderno_frecuencias_por_carta(df_mod)
    mod_frec_global = moderno_frecuencia_global(df_mod)
    mod_tfidf_carta = moderno_tfidf_por_carta(df_mod)
    mod_asociaciones_carta = moderno_asociaciones_por_carta(df_mod)
    mod_asociaciones_global_df = moderno_asociaciones_global(df_mod)
    mod_oraciones = moderno_longitud_oraciones(df_mod)

    # B) Comparación y reemplazos.
    print("\n[B] Cargando corpus original + moderno...")
    df_comp = cargar_corpus_comparacion(carpeta_original, carpeta_moderna)

    print(f"Páginas para comparación: {len(df_comp)}")
    print(f"Cartas para comparación: {df_comp['carta_id'].nunique()}")

    print("Calculando reemplazos original -> moderno...")
    reemplazos = analizar_reemplazos(df_comp)
    reemplazos_detalle = reemplazos.get("reemplazos_detalle", pd.DataFrame())
    reemplazos_global = reemplazos.get("reemplazos_global", pd.DataFrame())
    reemplazos_por_carta = reemplazos.get("reemplazos_por_carta", pd.DataFrame())
    reemplazos_por_pagina = reemplazos.get("reemplazos_por_pagina", pd.DataFrame())
    reemplazos_resumen_carta = reemplazos.get("reemplazos_resumen_carta", pd.DataFrame())
    reemplazos_resumen_pagina = reemplazos.get("reemplazos_resumen_pagina", pd.DataFrame())

    print("Calculando comparación...")
    comp_global = comparacion_global(df_comp, carpeta_original, carpeta_moderna)
    comp_pagina = comparacion_por_pagina(df_comp)
    comp_carta = comparacion_por_carta(df_comp)
    comp_cambios = cambios_frecuencia_por_carta(df_comp)
    comp_solo_original, comp_solo_moderno = vocabulario_exclusivo(df_comp)

    # Excel moderno.
    hojas_moderno = {
        "resumen_global": mod_global,
        "resumen_por_pagina": mod_pagina,
        "resumen_por_carta": mod_carta,
        "frecuencia_por_pagina": mod_frec_pagina,
        "frecuencia_por_carta": mod_frec_carta,
        "frecuencia_global": mod_frec_global,
        "tfidf_por_carta": mod_tfidf_carta,
        "asociaciones_por_carta": mod_asociaciones_carta,
        "asociaciones_global": mod_asociaciones_global_df.head(5000) if not mod_asociaciones_global_df.empty else mod_asociaciones_global_df,
        "longitud_oraciones": mod_oraciones,
        "reemplazos_global": reemplazos_global,
        "reemplazos_por_carta": reemplazos_por_carta,
        "reemplazos_por_pagina": reemplazos_por_pagina,
        "reemplazos_resumen_carta": reemplazos_resumen_carta,
        "reemplazos_resumen_pagina": reemplazos_resumen_pagina,
        "indice_archivos_modernos": df_mod.drop(columns=["texto_moderno"]),
    }

    print("Exportando análisis moderno en Excel...")
    exportar_excel(EXCEL_MODERNO, hojas_moderno)

    if not reemplazos_detalle.empty:
        exportar_excel_df_grande(
            CARPETA_REEMPLAZOS / "reemplazos_detalle_completo.xlsx",
            reemplazos_detalle,
            base_sheet="reemplazos_detalle"
        )

    # Excel comparación.
    hojas_comp = {
        "resumen_global": comp_global,
        "comparacion_por_pagina": comp_pagina,
        "comparacion_por_carta": comp_carta,
        "cambios_frecuencia_carta": comp_cambios,
        "vocab_solo_original": comp_solo_original,
        "vocab_solo_moderno": comp_solo_moderno,
        "reemplazos_global": reemplazos_global,
        "reemplazos_resumen_carta": reemplazos_resumen_carta,
        "reemplazos_resumen_pagina": reemplazos_resumen_pagina,
        "indice_archivos": df_comp.drop(columns=["texto_original", "texto_moderno"]),
    }

    print("Exportando comparación en Excel...")
    exportar_excel(EXCEL_COMPARACION, hojas_comp)

    print("Generando gráficos, redes, nubes y HTML...")
    graficos_moderno(mod_carta, mod_frec_global)
    graficos_reemplazos(reemplazos_global)
    graficos_comparacion(comp_carta, comp_cambios)
    generar_nubes_moderno(df_mod)
    generar_redes_asociaciones(mod_asociaciones_carta, mod_asociaciones_global_df)
    generar_red_reemplazos(reemplazos_global)
    generar_html_comparaciones(df_comp)

    guardar_resumenes(carpeta_original, carpeta_moderna)

    print("\nProceso terminado correctamente.")
    print("\nANÁLISIS SOLO MODERNO:")
    print(f"Excel: {EXCEL_MODERNO.resolve()}")
    print("CSV: no se genera ninguna carpeta ni archivo CSV.")
    print(f"Gráficos: {CARPETA_MODERNO_GRAFICOS.resolve()}")
    print(f"Redes: {CARPETA_MODERNO_REDES.resolve()}")
    print(f"Nubes: {CARPETA_MODERNO_NUBES.resolve()}")
    print(f"Reemplazos: {CARPETA_REEMPLAZOS.resolve()}")

    print("\nCOMPARACIÓN ORIGINAL VS MODERNO:")
    print(f"Excel: {EXCEL_COMPARACION.resolve()}")
    print("CSV: no se genera ninguna carpeta ni archivo CSV.")
    print(f"Gráficos: {CARPETA_COMP_GRAFICOS.resolve()}")
    print(f"HTML comparación: {CARPETA_COMP_HTML.resolve()}")


if __name__ == "__main__":
    main()
